"""
Smart Money Stock Tracker
==========================
Tracks top holdings and recent trades for four "smart money" portfolios:

    • ARK Innovation ETF  (ARKK) — Cathie Wood's fund, daily holdings
    • U.S. House Members  (e.g. Nancy Pelosi) — STOCK Act disclosures
    • Trump               — Presidential financial disclosures (annual)
    • Soros Fund Mgmt     — SEC 13F quarterly filings

Output: a single Excel file  →  output/smart_money_tracker.xlsx
  Sheet 1 "Top Holdings"    — current top positions, side-by-side
  Sheet 2 "Recent Activity" — individual trades (Congress) + QoQ changes (Soros)

Usage
-----
    python main.py                         # all defaults from .env / defaults
    python main.py --top 12                # show top 12 per column
    python main.py --members "Nancy Pelosi,Alexandria Ocasio-Cortez"
    python main.py --ark ARKW
    python main.py --days 365              # look back 1 year for trades
    python main.py --no-trump              # skip Trump column
    python main.py --no-soros             # skip Soros column
"""

import argparse
import logging
import os
import sys
from datetime import datetime
from pathlib import Path

# ── UTF-8 output on Windows ───────────────────────────────────────────────────
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

import pandas as pd
from dotenv import load_dotenv

_env_path = Path(__file__).parent / ".env"
load_dotenv(_env_path)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

from scrapers.ark      import get_ark_holdings
from scrapers.congress import get_congress_holdings, get_congress_trades
from scrapers.soros    import get_soros_holdings, get_soros_changes
from scrapers.trump    import get_trump_holdings, get_trump_trades


# ── Formatting helpers ────────────────────────────────────────────────────────

def _fmt_usd(val) -> str:
    """Compact, readable dollar string."""
    if val is None or val == "":
        return ""
    v = float(val)
    if abs(v) >= 1_000_000_000:
        return f"${v/1_000_000_000:.1f}B"
    if abs(v) >= 1_000_000:
        return f"${v/1_000_000:.1f}M"
    if abs(v) >= 1_000:
        return f"${v/1_000:.0f}K"
    return f"${v:.0f}"

def _fmt_pct(val) -> str:
    if val is None or val == "":
        return ""
    return f"{float(val):.2f}%"

def _pad(lst, length):
    empty = {"ticker": "", "company": "", "weight": None, "market_value": None}
    return lst + [empty] * max(0, length - len(lst))


# ── Sheet 1: Top Holdings ─────────────────────────────────────────────────────

def build_holdings_sheet(
    ark_data,
    congress_cols,  # [(label, holdings, date), ...]
    trump_data,
    soros_data,
    top_n,
    include_trump,
    include_soros,
) -> pd.DataFrame:
    """
    Wide-format table: one row per rank, two columns per portfolio
    (Ticker + Value/Weight).
    """
    ark_h, ark_date   = ark_data
    soros_h, sor_date = soros_data
    trump_h, trp_date = trump_data

    rows = []
    for i in range(top_n):
        row = {"Rank": i + 1}

        # ── ARK ──────────────────────────────────────────────────────────────
        a = _pad(ark_h, top_n)[i]
        row["ARK (ARKK) Ticker"] = a["ticker"]
        row["ARK Weight"]        = _fmt_pct(a["weight"])

        # ── Congress member(s) ───────────────────────────────────────────────
        for label, holdings, _ in congress_cols:
            col = label.split()[-1]          # "Pelosi", "Ocasio-Cortez", etc.
            c   = _pad(holdings, top_n)[i]
            row[f"{col} Ticker"]       = c["ticker"]
            row[f"{col} Net Purchased"] = _fmt_usd(c["market_value"])

        # ── Trump ─────────────────────────────────────────────────────────────
        if include_trump:
            t = _pad(trump_h, top_n)[i]
            row["Trump Ticker"]   = t["ticker"]
            row["Trump Est. Value"] = _fmt_usd(t["market_value"])

        # ── Soros ─────────────────────────────────────────────────────────────
        if include_soros:
            s = _pad(soros_h, top_n)[i]
            row["Soros Ticker"]  = s["ticker"]
            row["Soros Weight"]  = _fmt_pct(s["weight"])

        rows.append(row)

    df = pd.DataFrame(rows)

    # Append a blank separator and source-date footer
    footer_rows = [
        {col: "" for col in df.columns},
        {"Rank": "DATA SOURCES & DATES"},
    ]
    footer_rows[-1]["ARK (ARKK) Ticker"] = f"ARK: {ark_date or 'n/a'}"
    if include_trump:
        footer_rows[-1]["Trump Ticker"] = f"Trump: {trp_date or 'see trump_holdings.json'}"
    if include_soros:
        footer_rows[-1]["Soros Ticker"] = f"Soros 13F: {sor_date or 'n/a'}"
    for label, _, date in congress_cols:
        col = label.split()[-1]
        footer_rows[-1][f"{col} Ticker"] = f"{label}: {date or 'n/a'}"

    df = pd.concat([df, pd.DataFrame(footer_rows)], ignore_index=True)
    return df


# ── Sheet 2: Recent Activity ──────────────────────────────────────────────────

def build_activity_sheet(
    congress_cols_trades,  # [(label, trades_list, date), ...]
    trump_trades,          # list of trade dicts (from White House PTRs)
    soros_changes,
    soros_change_date,
    include_trump,
    include_soros,
) -> pd.DataFrame:
    """
    Long-format activity log combining:
      - Individual Congress member trades (from PTR filings)
      - Soros Fund quarter-over-quarter position changes

    Columns: Date | Portfolio | Ticker | Company | Action | Position Type |
             Amount / Value | Change vs Prior | Notes
    """
    rows = []

    # ── Trump PTR trades ──────────────────────────────────────────────────────
    if include_trump and trump_trades:
        for t in trump_trades:
            amount_str = t.get("raw_amount") or _fmt_usd(t.get("amount_mid"))
            pos_type   = "SHORT / Bearish" if not t.get("is_long", True) else "LONG"
            rows.append({
                "Date":           t.get("filing_date", ""),
                "Portfolio":      "Trump (White House PTR)",
                "Ticker":         t["ticker"],
                "Company":        t.get("company", ""),
                "Action":         t["action"],
                "Position Type":  pos_type,
                "Amount / Value": amount_str,
                "Change vs Prior": "",
                "Notes":          f"Instrument: {t['instrument']}  |  Source: White House PTR filing",
            })

    # ── Congress trades ───────────────────────────────────────────────────────
    for label, trades, _ in congress_cols_trades:
        for t in trades:
            amount_str = t.get("raw_amount") or _fmt_usd(t.get("amount_mid"))
            pos_type   = "SHORT / Bearish" if not t.get("is_long", True) else "LONG"
            rows.append({
                "Date":              t["filing_date"],
                "Portfolio":         label,
                "Ticker":            t["ticker"],
                "Company":           t.get("company", ""),
                "Action":            t["action"],
                "Position Type":     pos_type,
                "Amount / Value":    amount_str,
                "Change vs Prior":   "",
                "Notes":             f"Instrument: {t['instrument']}  |  Source: House PTR filing",
            })

    # ── Soros changes ─────────────────────────────────────────────────────────
    if include_soros and soros_changes:
        sample = soros_changes[0] if soros_changes else {}
        cur_p  = sample.get("current_period", "")
        prev_p = sample.get("previous_period", "")
        note_pfx = f"Soros 13F: {cur_p} vs {prev_p}"

        for c in soros_changes:
            if c["change_type"] == "Closed":
                action = "CLOSED position"
            elif c["change_type"] == "New Position":
                action = "NEW position"
            elif c["change_type"] == "Increased":
                action = "INCREASED position"
            else:
                action = "REDUCED position"

            val_str    = _fmt_usd(c["current_value"]) if c["current_value"] else ""
            change_str = ""
            if c.get("pct_change") is not None:
                sign = "+" if c["value_change"] >= 0 else ""
                change_str = f"{sign}{c['pct_change']:.1f}%  ({_fmt_usd(c['value_change'])})"

            rows.append({
                "Date":              c["current_period"],
                "Portfolio":         "Soros Fund Mgmt",
                "Ticker":            c["ticker"],
                "Company":           c["company"],
                "Action":            action,
                "Position Type":     "LONG  (13F reports long equity only)",
                "Amount / Value":    val_str,
                "Change vs Prior":   change_str,
                "Notes":             note_pfx,
            })

    if not rows:
        rows = [{"Date": "No data available", "Portfolio": "", "Ticker": "",
                 "Company": "", "Action": "", "Position Type": "",
                 "Amount / Value": "", "Change vs Prior": "", "Notes": ""}]

    df = pd.DataFrame(rows)
    # Sort by date descending so most-recent is at the top
    df = df.sort_values("Date", ascending=False).reset_index(drop=True)
    return df


# ── Excel formatting ──────────────────────────────────────────────────────────

def _apply_excel_formatting(writer, sheet_name, df, freeze_row=1):
    """Apply column widths, bold headers, and row colours to an Excel sheet."""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return  # formatting is optional

    wb = writer.book
    ws = writer.sheets[sheet_name]
    ncols = len(df.columns)

    # ── Header row ────────────────────────────────────────────────────────────
    HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")   # dark navy
    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
    CENTRE       = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTRE

    ws.row_dimensions[1].height = 30

    # ── Row colours (alternating + action highlights) ─────────────────────────
    LIGHT_ROW   = PatternFill("solid", fgColor="EBF3FB")   # pale blue
    WHITE_ROW   = PatternFill("solid", fgColor="FFFFFF")
    BUY_FILL    = PatternFill("solid", fgColor="E2EFDA")   # light green
    SELL_FILL   = PatternFill("solid", fgColor="FCE4D6")   # light red/orange
    NEW_FILL    = PatternFill("solid", fgColor="FFF2CC")   # light yellow
    FOOTER_FILL = PatternFill("solid", fgColor="D9D9D9")   # grey

    action_col = None
    for j, col_name in enumerate(df.columns, 1):
        if col_name == "Action":
            action_col = j
            break

    for row_idx, (_, row_data) in enumerate(df.iterrows(), start=2):
        action_val = str(row_data.get("Action", "")).upper() if action_col else ""
        rank_val   = str(row_data.get("Rank", "")).upper()

        if rank_val == "DATA SOURCES & DATES":
            fill = FOOTER_FILL
        elif "BOUGHT" in action_val or "NEW" in action_val or "INCREASED" in action_val:
            fill = BUY_FILL
        elif "SOLD" in action_val or "REDUCED" in action_val or "CLOSED" in action_val:
            fill = SELL_FILL
        elif row_idx % 2 == 0:
            fill = LIGHT_ROW
        else:
            fill = WHITE_ROW

        for col_idx in range(1, ncols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")

    # ── Column widths ─────────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_idx)
        # Measure content width
        max_len = max(
            len(str(col_name)),
            df.iloc[:, col_idx - 1].astype(str).str.len().max() if len(df) > 0 else 0,
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 45)

    # ── Freeze header row ─────────────────────────────────────────────────────
    ws.freeze_panes = f"A{freeze_row + 1}"

    # ── Thin border on header cells ───────────────────────────────────────────
    thin = Side(style="thin", color="AAAAAA")
    for col_idx in range(1, ncols + 1):
        ws.cell(row=1, column=col_idx).border = Border(
            bottom=Side(style="medium", color="FFFFFF")
        )


def _write_notes_sheet(writer, include_trump, include_soros):
    """Add a 'Notes & Limitations' sheet explaining each data source."""
    notes = [
        ["SMART MONEY STOCK TRACKER — NOTES & LIMITATIONS", ""],
        ["", ""],
        ["WHAT IS BEING TRACKED?", ""],
        ["ARK Innovation ETF (ARKK)",
         "Daily holdings snapshot of Cathie Wood's flagship fund. "
         "Shows current portfolio weights. ARK is a long-only fund — no short positions. "
         "Source: arkfunds.io (unofficial aggregator of ARK's published daily holdings CSV)."],
        ["Congress / House Members",
         "Individual stock TRANSACTIONS disclosed under the STOCK Act "
         "(Periodic Transaction Reports, or PTRs). Politicians must disclose "
         "trades within 45 days. Shows buys and sells in dollar ranges. "
         "'Net Purchased' = sum of purchases minus sales in the look-back window. "
         "Source: disclosures-clerk.house.gov (official House of Representatives portal)."],
        ["Trump",
         "Two automated sources combined:\n"
         "1) DJT (Trump Media) stake: fetched live from SEC EDGAR Form 4 filings "
         "(Trump's Trust holds ~114.75M shares; multiplied by live Yahoo Finance price). "
         "This is the most accurate and up-to-date figure.\n"
         "2) Other holdings (AAPL, AMZN, NVDA, etc.): from the annual OGE Form 278e "
         "presidential financial disclosure in data/trump_holdings.json. "
         "Filed once per year — may lag by up to 12 months.\n"
         "3) Trades (Recent Activity sheet): from White House Periodic Transaction Reports "
         "(PTRs, the same STOCK Act disclosures required of Congress). "
         "PDFs are image-scanned; set OPENROUTER_API_KEY in .env to enable LLM parsing. "
         "Source: whitehouse.gov/disclosures/ + SEC EDGAR (sec.gov)."],
        ["Soros Fund Management",
         "Quarterly 13F-HR filing with the SEC. Shows long equity positions only. "
         "Filed within 45 days of quarter-end. Source: SEC EDGAR (official)."],
        ["", ""],
        ["LONG vs SHORT POSITIONS", ""],
        ["What is shown",
         "All data sources above primarily show LONG positions (bets that a stock rises)."],
        ["Congress — Options",
         "Put options (bearish bets) ARE flagged in the 'Recent Activity' sheet "
         "under 'Position Type = SHORT / Bearish'. "
         "These are extracted from the PTR filing descriptions."],
        ["Soros — Short positions",
         "13F filings by law only require disclosure of LONG equity positions. "
         "Short positions are NOT included in 13F data. "
         "Soros's short bets (if any) are not captured here — this is a legal/structural "
         "limitation of the SEC filing system, not a scraping limitation."],
        ["ARK — Short positions",
         "ARK funds do not take short positions by their investment mandate."],
        ["Trump — Short positions",
         "Presidential disclosures list asset holdings; short positions are rarely disclosed "
         "unless specifically held. Not captured here."],
        ["", ""],
        ["DATA FRESHNESS", ""],
        ["ARK",          "Updated daily (reflects yesterday's close)."],
        ["Congress",     "Depends on member filing promptly. Law requires filing within 45 days of trade."],
        ["Soros",        "Quarterly — typically 6-8 weeks behind real-time (e.g. Q4 data available Feb)."],
        ["Trump",
         "DJT stake: real-time (SEC Form 4 + live price feed). "
         "Other holdings: annual (OGE Form 278e, may lag 12 months). "
         "PTR trades: requires OPENROUTER_API_KEY; filed within 30 days per STOCK Act."],
        ["", ""],
        ["DISCLAIMER",
         "This tool aggregates publicly available government disclosures for educational purposes. "
         "Nothing here is investment advice. Past trades by public figures do not predict future returns."],
    ]

    df_notes = pd.DataFrame(notes, columns=["Topic", "Explanation"])
    df_notes.to_excel(writer, sheet_name="Notes & Sources", index=False)

    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        ws = writer.sheets["Notes & Sources"]
        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 90

        TITLE_FONT    = Font(bold=True, size=13, color="1F4E79")
        HEADER_FONT   = Font(bold=True, size=11)
        SECTION_FILL  = PatternFill("solid", fgColor="D6E4F0")
        WRAP          = Alignment(vertical="top", wrap_text=True)

        for row in ws.iter_rows():
            topic_cell = row[0]
            if topic_cell.value in ("SMART MONEY STOCK TRACKER — NOTES & LIMITATIONS",):
                topic_cell.font = TITLE_FONT
            elif topic_cell.value and topic_cell.value.isupper() and len(topic_cell.value) > 3:
                topic_cell.font = HEADER_FONT
                for cell in row:
                    cell.fill = SECTION_FILL
            for cell in row:
                cell.alignment = WRAP
            ws.row_dimensions[row[0].row].height = None   # auto-height

        ws.freeze_panes = "A2"
    except Exception:
        pass


# ── Main ──────────────────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(description="Smart Money Stock Tracker")
    p.add_argument("--top",     type=int,  default=int(os.getenv("TOP_N", 10)))
    p.add_argument("--members", type=str,  default=os.getenv("CONGRESS_MEMBERS", "Nancy Pelosi"))
    p.add_argument("--ark",     type=str,  default=os.getenv("ARK_FUND", "ARKK"))
    p.add_argument("--days",    type=int,  default=int(os.getenv("CONGRESS_DAYS_BACK", 365)))
    p.add_argument("--output",  type=str,  default=os.getenv("OUTPUT_DIR", "output"))
    p.add_argument("--no-trump",  action="store_true")
    p.add_argument("--no-soros",  action="store_true")
    p.add_argument("--no-ark",    action="store_true")
    return p.parse_args()


def main():
    args    = parse_args()
    top_n   = args.top
    members = [m.strip() for m in args.members.split(",") if m.strip()]

    print()
    print("=" * 68)
    print("  Smart Money Stock Tracker")
    print(f"  Generated: {datetime.now().strftime('%Y-%m-%d  %H:%M')}")
    print("=" * 68)

    # ── Fetch Holdings ────────────────────────────────────────────────────────
    ark_data = ([], None)
    if not args.no_ark:
        print(f"\n[1/5] ARK {args.ark} current holdings ...")
        ark_data = get_ark_holdings(args.ark, top_n)

    congress_holdings_cols = []
    print(f"\n[2/5] Congressional top holdings (net purchases) ...")
    for member in members:
        print(f"       {member} ...")
        h, date = get_congress_holdings(member, top_n, args.days)
        congress_holdings_cols.append((member, h, date))

    trump_data = ([], None)
    if not args.no_trump:
        print("\n[3/5] Trump financial disclosure holdings ...")
        trump_data = get_trump_holdings(top_n)

    soros_holdings_data = ([], None)
    soros_changes       = []
    soros_change_date   = None
    if not args.no_soros:
        print("\n[4/5] Soros Fund Management 13F — current holdings ...")
        soros_holdings_data = get_soros_holdings(top_n)

        print("\n[5/5] Soros Fund Management 13F — quarter-over-quarter changes ...")
        soros_changes, soros_change_date = get_soros_changes(top_n)

    # ── Fetch Individual Trades ───────────────────────────────────────────────
    congress_trades_cols = []
    print("\n[+] Fetching individual congressional trade records ...")
    for member in members:
        print(f"       {member} ...")
        trades, date = get_congress_trades(member, args.days)
        congress_trades_cols.append((member, trades, date))

    trump_trades_list: list = []
    if not args.no_trump:
        print("\n[+] Fetching Trump White House PTR trade records ...")
        trump_trades_list, _ = get_trump_trades(args.days)
        if trump_trades_list:
            print(f"       {len(trump_trades_list)} trades found from White House PTRs")
        else:
            print("       No trades found (set OPENROUTER_API_KEY to enable PDF parsing)")

    # ── Build DataFrames ──────────────────────────────────────────────────────
    holdings_df = build_holdings_sheet(
        ark_data, congress_holdings_cols, trump_data, soros_holdings_data,
        top_n, not args.no_trump, not args.no_soros,
    )
    activity_df = build_activity_sheet(
        congress_trades_cols, trump_trades_list,
        soros_changes, soros_change_date,
        not args.no_trump, not args.no_soros,
    )

    # ── Write Excel ───────────────────────────────────────────────────────────
    out_dir = Path(args.output)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "smart_money_tracker.xlsx"

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        holdings_df.to_excel(writer, sheet_name="Top Holdings",     index=False)
        activity_df.to_excel(writer, sheet_name="Recent Activity",  index=False)
        _write_notes_sheet(writer, not args.no_trump, not args.no_soros)

        _apply_excel_formatting(writer, "Top Holdings",    holdings_df)
        _apply_excel_formatting(writer, "Recent Activity", activity_df)

    # ── Summary ───────────────────────────────────────────────────────────────
    print(f"\n{'─' * 68}")
    print(f"  Output: {out_path.resolve()}")
    print(f"  Sheets: Top Holdings | Recent Activity | Notes & Sources")
    print(f"{'─' * 68}")
    print(f"\n  Top Holdings summary:")
    print(f"    ARK {args.ark}      : {len(ark_data[0])} holdings  (as of {ark_data[1] or 'n/a'})")
    for label, h, date in congress_holdings_cols:
        print(f"    {label:<15}: {len(h)} net-buy positions  (latest PTR: {date or 'n/a'})")
    if not args.no_trump:
        print(f"    Trump          : {len(trump_data[0])} disclosed holdings  ({trump_data[1] or 'see JSON file'})")
    if not args.no_soros:
        print(f"    Soros 13F      : {len(soros_holdings_data[0])} top holdings  (period: {soros_holdings_data[1] or 'n/a'})")
    print(f"\n  Recent Activity summary:")
    for label, trades, _ in congress_trades_cols:
        print(f"    {label:<15}: {len(trades)} individual trades")
    if not args.no_soros:
        print(f"    Soros changes  : {len(soros_changes)} position changes")
    print()

    print("  LIMITATIONS:")
    print("  * Short positions: Congress put options are flagged; Soros shorts")
    print("    NOT available (13F only covers long equity by law).")
    print("  * Trump data is annual (OGE Form 278e), not real-time.")
    print("  * Pelosi 'Net Purchased' = estimated net from disclosed trades only.")
    print()


if __name__ == "__main__":
    main()
